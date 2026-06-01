"""Unit tests for the Damodaran importer (Block 2 hardening).

Covers:
- Fix 1: download failure is caught, logged, and returned as an error IngestResult.
- Fix 2: atomic dual upsert — a failure in the second upsert rolls back the first.
- Fix 3: empty-row import yields a "partial" IngestResult with a descriptive message.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any
from unittest.mock import patch

import httpx
import openpyxl

from bot.ingest.damodaran import import_damodaran, import_damodaran_from_files
from bot.storage.db import apply_schema, connect

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_industry_xlsx(path: Path) -> None:
    """Write a minimal industry xlsx that the importer can parse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Industry Averages"
    ws.append(
        [
            "Industry Name",
            "Beta",
            "Cost of Equity",
            "Cost of Debt",
            "Tax Rate",
            "Cost of Capital",
        ]
    )
    ws.append(["Software", 1.1, 0.10, 0.04, 0.25, 0.09])
    ws.append(["Retail", 0.9, 0.08, 0.035, 0.22, 0.07])
    wb.save(path)


def _make_country_xlsx(path: Path) -> None:
    """Write a minimal country xlsx that the importer can parse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERPs by country"
    ws.append(
        [
            "Country",
            "Africa",
            "Moody's rating",
            "Total Equity Risk Premium",
            "Country Risk Premium",
        ]
    )
    ws.append(["United States", "North America", "Aaa", 0.05, 0.0])
    ws.append(["Brazil", "Latin America", "Ba2", 0.09, 0.04])
    wb.save(path)


def _make_conn() -> Any:
    conn = connect(":memory:")
    apply_schema(conn)
    return conn


# ---------------------------------------------------------------------------
# Fix 1 — download failure → error IngestResult
# ---------------------------------------------------------------------------


def test_import_damodaran_download_failure_returns_error_result(tmp_path: Path) -> None:
    """If download_dataset raises, import_damodaran returns an error IngestResult
    and records the failure in refresh_log."""
    conn = _make_conn()

    with patch(
        "bot.ingest.damodaran.download_dataset",
        side_effect=httpx.HTTPError("connection refused"),
    ):
        result = import_damodaran(
            conn,
            download_dir=tmp_path,
            region="US",
            year=2026,
        )

    assert result.status == "error"
    assert result.error_message is not None
    assert "download failed" in result.error_message

    rows = conn.execute(
        "SELECT status, error_message FROM refresh_log WHERE source = 'damodaran'"
    ).fetchall()
    assert len(rows) == 1
    assert rows[0][0] == "error"
    assert rows[0][1] is not None and "download failed" in rows[0][1]

    conn.close()


def test_import_damodaran_success_writes_single_refresh_log_row(
    tmp_path: Path,
) -> None:
    """The success path must write exactly one 'damodaran' refresh_log row.

    Regression: import_damodaran wrapped the download in its own refresh_run
    envelope AND delegated to import_damodaran_from_files (which opens a second
    envelope), so a successful run logged two rows — a phantom rows_affected=0
    'success' plus the real row.
    """
    conn = _make_conn()

    industry_xlsx = tmp_path / "wacc.xlsx"
    country_xlsx = tmp_path / "ctryprem.xlsx"
    _make_industry_xlsx(industry_xlsx)
    _make_country_xlsx(country_xlsx)

    # download_dataset just returns the path it was asked to write to; the files
    # already exist on disk so the importer can parse them.
    def _fake_download(url: str, dest: Path) -> Path:
        return industry_xlsx if "wacc" in dest.name else country_xlsx

    with patch("bot.ingest.damodaran.download_dataset", side_effect=_fake_download):
        result = import_damodaran(
            conn,
            download_dir=tmp_path,
            region="US",
            year=2026,
        )

    assert result.status == "success"
    assert result.rows_affected > 0

    rows = conn.execute(
        "SELECT status, rows_affected FROM refresh_log WHERE source = 'damodaran'"
    ).fetchall()
    assert len(rows) == 1, f"expected exactly one refresh_log row, got {rows}"
    assert rows[0][0] == "success"
    assert rows[0][1] == result.rows_affected

    conn.close()


def test_import_damodaran_import_failure_writes_single_error_row(
    tmp_path: Path,
) -> None:
    """A parse/upsert failure after a successful download must still produce a
    single error refresh_log row (one envelope, not two)."""
    conn = _make_conn()

    industry_xlsx = tmp_path / "wacc.xlsx"
    country_xlsx = tmp_path / "ctryprem.xlsx"
    _make_industry_xlsx(industry_xlsx)
    _make_country_xlsx(country_xlsx)

    def _fake_download(url: str, dest: Path) -> Path:
        return industry_xlsx if "wacc" in dest.name else country_xlsx

    with (
        patch("bot.ingest.damodaran.download_dataset", side_effect=_fake_download),
        patch(
            "bot.ingest.damodaran.upsert_industry_rows",
            side_effect=RuntimeError("boom"),
        ),
    ):
        result = import_damodaran(
            conn,
            download_dir=tmp_path,
            region="US",
            year=2026,
        )

    assert result.status == "error"
    assert result.error_message is not None and "boom" in result.error_message

    rows = conn.execute("SELECT status FROM refresh_log WHERE source = 'damodaran'").fetchall()
    assert len(rows) == 1, f"expected exactly one refresh_log row, got {rows}"
    assert rows[0][0] == "error"

    conn.close()


# ---------------------------------------------------------------------------
# Fix 2 — atomic dual upsert (rollback on second-upsert failure)
# ---------------------------------------------------------------------------


def test_import_damodaran_from_files_rolls_back_on_second_upsert_failure(
    tmp_path: Path,
) -> None:
    """If the country upsert fails, the industry rows inserted in the same
    transaction must be rolled back (industry table stays empty)."""
    conn = _make_conn()

    industry_xlsx = tmp_path / "wacc.xlsx"
    country_xlsx = tmp_path / "ctry.xlsx"
    _make_industry_xlsx(industry_xlsx)
    _make_country_xlsx(country_xlsx)

    with patch(
        "bot.ingest.damodaran.upsert_country_rows",
        side_effect=RuntimeError("simulated country upsert failure"),
    ):
        result = import_damodaran_from_files(
            conn,
            industry_path=industry_xlsx,
            country_path=country_xlsx,
            region="US",
            year=2026,
        )

    assert result.status == "error"

    # The industry rows written inside the same transaction must be gone.
    industry_count = conn.execute("SELECT COUNT(*) FROM damodaran_industry").fetchone()[0]
    assert industry_count == 0, (
        f"Industry rows should have been rolled back, but found {industry_count}"
    )

    conn.close()


# ---------------------------------------------------------------------------
# Fix 3 — empty-row import → "partial" IngestResult
# ---------------------------------------------------------------------------


def test_import_damodaran_from_files_partial_when_industry_empty(
    tmp_path: Path,
) -> None:
    """If the industry file yields no rows, status should be 'partial'."""
    conn = _make_conn()

    # A country file with real rows.
    country_xlsx = tmp_path / "ctry.xlsx"
    _make_country_xlsx(country_xlsx)

    # An industry file that produces no data rows (header-only).
    empty_xlsx = tmp_path / "empty_wacc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Industry Averages"
    ws.append(
        ["Industry Name", "Beta", "Cost of Equity", "Cost of Debt", "Tax Rate", "Cost of Capital"]
    )
    wb.save(empty_xlsx)

    result = import_damodaran_from_files(
        conn,
        industry_path=empty_xlsx,
        country_path=country_xlsx,
        region="US",
        year=2026,
    )

    assert result.status == "partial"
    assert result.error_message is not None
    assert "industry" in result.error_message

    # Country rows should still have been written.
    country_count = conn.execute("SELECT COUNT(*) FROM damodaran_country").fetchone()[0]
    assert country_count > 0

    conn.close()


def test_import_damodaran_from_files_partial_when_country_empty(
    tmp_path: Path,
) -> None:
    """If the country file yields no rows, status should be 'partial'."""
    conn = _make_conn()

    industry_xlsx = tmp_path / "wacc.xlsx"
    _make_industry_xlsx(industry_xlsx)

    # A country file that produces no data rows (header-only).
    empty_xlsx = tmp_path / "empty_ctry.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERPs by country"
    ws.append(
        ["Country", "Africa", "Moody's rating", "Total Equity Risk Premium", "Country Risk Premium"]
    )
    wb.save(empty_xlsx)

    result = import_damodaran_from_files(
        conn,
        industry_path=industry_xlsx,
        country_path=empty_xlsx,
        region="US",
        year=2026,
    )

    assert result.status == "partial"
    assert result.error_message is not None
    assert "country" in result.error_message

    # Industry rows should still have been written.
    industry_count = conn.execute("SELECT COUNT(*) FROM damodaran_industry").fetchone()[0]
    assert industry_count > 0

    conn.close()
