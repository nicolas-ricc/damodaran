"""Damodaran datasets — parser (Capa A).

Column headers reflect the January 2026 Damodaran release.  If headers drift in
a future release, update ``DEFAULT_INDUSTRY_COLUMN_MAP`` and
``DEFAULT_COUNTRY_COLUMN_MAP`` to match the new strings; no other code change is
required.
"""

from __future__ import annotations

import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb
import httpx

from bot.ingest.base import IngestResult, RefreshOutcome, refresh_run, transaction
from bot.utils.logging import get_logger

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Column maps — db_column_name -> xls_header_string
# Inspected from Damodaran Jan-2026 files:
#   wacc.xls      sheet "Industry Averages"   header row 18
#   ctryprem.xls  sheet "ERPs by country"     header row 7
# ---------------------------------------------------------------------------

DEFAULT_INDUSTRY_COLUMN_MAP: dict[str, str] = {
    "industry": "Industry Name",
    "beta_levered": "Beta",
    "cost_of_equity": "Cost of Equity",
    "cost_of_debt": "Cost of Debt",
    "tax_rate": "Tax Rate",
    # "Cost of Capital" is labelled "WACC" in the DB schema
    "wacc": "Cost of Capital",
    # Parsing artefact, not a DB column: the published file carries D/(D+E) and
    # `derive_industry_columns` turns it into debt_to_equity + beta_unlevered.
    "debt_weight_raw": "D/(D+E)",
}

DEFAULT_COUNTRY_COLUMN_MAP: dict[str, str] = {
    "country": "Country",
    # The region column header is literally "Africa" in the file (column B),
    # but the values are regional groupings such as "Middle East", "Western
    # Europe", etc.  We capture it under "region" in the DB.
    "region": "Africa",
    "rating": "Moody's rating",
    "erp": "Total Equity Risk Premium",
    "country_risk_premium": "Country Risk Premium",
}

_DATASET_BASE = "https://pages.stern.nyu.edu/~adamodar/pc/datasets/"

INDUSTRY_WACC_URL = _DATASET_BASE + "wacc.xls"
COUNTRY_RISK_URL = _DATASET_BASE + "ctryprem.xls"


@dataclass(frozen=True)
class IndustryDataset:
    """One published Damodaran industry file and how to read it.

    Attributes:
        key: Short identifier, also the fixture filename stem and the key callers
            use in ``extra_industry_paths``.
        url: Published location.
        sheet_keywords: Passed to the existing sheet auto-detection.
        column_map: DB column → exact header string, as observed in the file.
            ``industry`` must be first: ``_to_normalized_rows`` treats the first
            key as the primary key of the row.
    """

    key: str
    url: str
    sheet_keywords: tuple[str, ...]
    column_map: dict[str, str]


#: Industry datasets merged into ``damodaran_industry``, in precedence order: an
#: earlier dataset's value for a column wins over a later one's.
#:
#: Sheet names and header strings below were observed directly from the published
#: files (see the discovery step in the plan) — do NOT guess them: a wrong header
#: silently yields a NULL column, which is the failure mode this registry exists
#: to eliminate.
INDUSTRY_DATASETS: tuple[IndustryDataset, ...] = (
    IndustryDataset(
        key="wacc",
        url=INDUSTRY_WACC_URL,
        sheet_keywords=("industry", "average"),
        column_map=DEFAULT_INDUSTRY_COLUMN_MAP,
    ),
    # margin.xls — margins. "Pre-tax Unadjusted Operating Margin" is the plain
    # operating margin; the file also publishes lease- and R&D-adjusted variants.
    IndustryDataset(
        key="margin",
        url=_DATASET_BASE + "margin.xls",
        sheet_keywords=("industry", "average"),
        column_map={
            "industry": "Industry Name",
            "op_margin": "Pre-tax Unadjusted Operating Margin",
            "net_margin": "Net Margin",
        },
    ),
    # capex.xls — sales-to-capital. Note the space after the slash in the header.
    # `reinvestment_rate` has no clean column here (only "Net Cap Ex/ EBIT (1-t)",
    # a different quantity) and no consumer SELECTs it, so it is left unmapped.
    IndustryDataset(
        key="capex",
        url=_DATASET_BASE + "capex.xls",
        sheet_keywords=("industry", "average"),
        column_map={
            "industry": "Industry Name",
            "sales_to_capital": "Sales/ Invested Capital (LTM)",
        },
    ),
    # pedata.xls — earnings multiples. "Current PE" is the trailing-12-month figure
    # the §7.7 sanity check compares against; Trailing/Forward PE are alternatives.
    IndustryDataset(
        key="pedata",
        url=_DATASET_BASE + "pedata.xls",
        sheet_keywords=("industry", "average"),
        column_map={"industry": "Industry Name", "pe": "Current PE"},
    ),
    # pbvdata.xls — book multiples plus both returns. ROIC lives here, not in
    # eva.xls (which 404s).
    IndustryDataset(
        key="pbvdata",
        url=_DATASET_BASE + "pbvdata.xls",
        sheet_keywords=("industry", "average"),
        column_map={
            "industry": "Industry Name",
            "pbv": "PBV",
            "roe": "ROE",
            "roic": "ROIC",
        },
    ),
    # vebitda.xls — EV/EBITDA. The sheet publishes EV/EBITDA twice; the record
    # loader suffixes the second one "_1", so this header takes the first.
    IndustryDataset(
        key="vebitda",
        url=_DATASET_BASE + "vebitda.xls",
        sheet_keywords=("industry", "average"),
        column_map={"industry": "Industry Name", "ev_ebitda": "EV/EBITDA"},
    ),
    # psdata.xls — EV/Sales, which vebitda.xls does not carry. Its Net Margin and
    # Pre-tax Operating Margin duplicate margin.xls; the merge gives margin.xls
    # precedence, so the overlap is harmless.
    IndustryDataset(
        key="psdata",
        url=_DATASET_BASE + "psdata.xls",
        sheet_keywords=("industry", "average"),
        column_map={"industry": "Industry Name", "ev_sales": "EV/Sales"},
    ),
)

#: The registry entry ``industry_path`` / ``industry_url`` refer to; the rest are
#: the "extra" datasets keyed by ``IndustryDataset.key``.
_WACC_DATASET = INDUSTRY_DATASETS[0]
_EXTRA_DATASETS: tuple[IndustryDataset, ...] = INDUSTRY_DATASETS[1:]


def merge_industry_datasets(
    parts: Sequence[list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    """Outer-join per-dataset industry rows on the ``industry`` label.

    Earlier datasets win: a column already carrying a non-``None`` value is not
    overwritten by a later dataset, so the cost-of-capital file stays authoritative
    for the columns it publishes. Rows without an ``industry`` are dropped. Insertion
    order of first appearance is preserved so the output is deterministic.
    """
    merged: dict[str, dict[str, Any]] = {}
    for part in parts:
        for row in part:
            industry = row.get("industry")
            if not isinstance(industry, str) or not industry.strip():
                continue
            target = merged.setdefault(industry, {"industry": industry})
            for column, value in row.items():
                if column == "industry" or value is None:
                    continue
                if target.get(column) is None:
                    target[column] = value
    return list(merged.values())

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

# Keywords checked in priority order (most-specific first).
# For country files, "erps by" matches "ERPs by country" before the more
# generic "country" keyword would pick up "Country Lookup".
_INDUSTRY_KEYWORDS = ("industry",)
_COUNTRY_KEYWORDS = ("erps by", "erp by", "erp")
_AVERAGE_KEYWORDS = ("average",)


def _find_header_row(rows: list[list[Any]]) -> int:
    """Return the index of the best column-header row in the first 25 rows.

    The strategy is to find the row that is:
    1. Entirely composed of string values (no numerics in non-empty cells), AND
    2. Has the most non-empty cells — the widest all-string row is the header.

    Damodaran sheets typically have a few preamble rows with 2-3 string cells
    followed by the true header which spans all columns.  Choosing the widest
    all-string row reliably picks the real header over preamble rows.
    """
    best_idx = 0
    best_count = 0
    for i, r in enumerate(rows[:25]):
        non_empty = [c for c in r if c is not None and str(c).strip() != ""]
        if len(non_empty) < 3:
            continue
        numeric_cells = [c for c in non_empty if isinstance(c, (int, float))]
        string_cells = [c for c in non_empty if isinstance(c, str)]
        # Only consider rows that are entirely strings (no numerics).
        if len(numeric_cells) == 0 and len(string_cells) > best_count:
            best_count = len(string_cells)
            best_idx = i
    return best_idx


def _pick_sheet(names: list[str], keywords: tuple[str, ...]) -> str:
    """Return the best-matching sheet name from *names* using *keywords*.

    Keywords are checked in priority order: the first keyword that matches any
    sheet wins, and among sheets matching that keyword the first one is returned.
    Falls back to the first sheet if nothing matches.
    """
    lower_names = [n.lower() for n in names]
    for kw in keywords:
        for i, lower in enumerate(lower_names):
            if kw in lower:
                return names[i]
    return names[0]


def _load_rows(
    path: Path,
    sheet_name: str | None,
    keywords: tuple[str, ...] | None = None,
) -> tuple[list[list[Any]], str]:
    """Load all cell values from *path*, returning (rows, picked_sheet_name).

    Tries openpyxl first (xlsx), then falls back to xlrd (legacy xls).

    *keywords* overrides the sheet-detection keywords used when *sheet_name* is
    ``None``; the default order matches every file the module read before the
    per-dataset registry existed.
    """
    import openpyxl
    import openpyxl.utils.exceptions

    detect = keywords or (_INDUSTRY_KEYWORDS + _COUNTRY_KEYWORDS + _AVERAGE_KEYWORDS)

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        names: list[str] = list(wb.sheetnames)
        if sheet_name is not None:
            if sheet_name not in names:
                raise ValueError(f"Sheet {sheet_name!r} not found in {path}. Available: {names}")
            picked = sheet_name
        else:
            picked = _pick_sheet(names, detect)
        ws = wb[picked]
        rows: list[list[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows, picked
    except ValueError:
        raise  # explicit sheet-not-found error — do not fall through to xlrd
    except (openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile, OSError) as exc:
        log.info(
            "damodaran.openpyxl_failed_falling_back_to_xlrd",
            path=str(path),
            error=str(exc),
        )

    import xlrd  # type: ignore[import-untyped]

    book = xlrd.open_workbook(str(path))
    names_xlrd: list[str] = book.sheet_names()
    if sheet_name is not None:
        if sheet_name not in names_xlrd:
            raise ValueError(f"Sheet {sheet_name!r} not found in {path}. Available: {names_xlrd}")
        picked_xlrd = sheet_name
    else:
        picked_xlrd = _pick_sheet(names_xlrd, detect)
    ws_xlrd = book.sheet_by_name(picked_xlrd)
    xlrd_rows: list[list[Any]] = [ws_xlrd.row_values(r) for r in range(ws_xlrd.nrows)]
    return xlrd_rows, picked_xlrd


def _load_to_records(
    path: Path,
    sheet_name: str | None,
    keywords: tuple[str, ...] | None = None,
) -> list[dict[str, Any]]:
    """Read xls/xlsx, detect the header row, and return a list of row dicts.

    Using raw dicts instead of a Polars DataFrame avoids type-inference issues
    with Damodaran sheets that mix strings, floats, and ``None`` in the same
    column.
    """
    rows, picked = _load_rows(path, sheet_name, keywords)
    log.debug("damodaran.sheet_picked", path=str(path), sheet=picked)

    if not rows:
        return []

    header_idx = _find_header_row(rows)
    header_raw = rows[header_idx]

    # Deduplicate and sanitize column names.
    seen: dict[str, int] = {}
    header: list[str] = []
    for i, h in enumerate(header_raw):
        name = str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        header.append(name)

    records: list[dict[str, Any]] = []
    for r in rows[header_idx + 1 :]:
        records.append({h: (r[i] if i < len(r) else None) for i, h in enumerate(header)})
    return records


# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------


def _coerce_value(value: Any) -> Any:
    """Coerce a single cell value: strip strings, parse percentage strings, coerce numeric strings."""
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        if value.endswith("%"):
            try:
                return float(value[:-1]) / 100.0
            except ValueError:
                pass
        # Try to coerce plain numeric strings (e.g. "1.23") to float.
        try:
            return float(value)
        except ValueError:
            pass
    return value


def _null_non_numeric_industry_cells(
    row: dict[str, Any],
    column_map: dict[str, str],
    path: Path,
) -> dict[str, Any]:
    """NULL out industry cells that are text where a number is expected.

    Every mapped ``damodaran_industry`` column except the industry label itself is
    a ``DOUBLE``, but the published sheets write ``"NA"`` (and the occasional other
    marker) where an industry has no meaningful value — e.g. a negative-earnings
    industry's PE. Passing that string to DuckDB raises a conversion error that
    would abort the whole import, so it becomes a NULL, which is what it means.
    """
    out = dict(row)
    pk_field = next(iter(column_map))
    for db_col in column_map:
        if db_col == pk_field:
            continue
        value = out.get(db_col)
        if value is not None and not isinstance(value, (int, float)):
            log.debug(
                "damodaran.industry.non_numeric_cell",
                path=str(path),
                column=db_col,
                industry=out.get(pk_field),
                value=str(value)[:40],
            )
            out[db_col] = None
    return out


def _to_normalized_rows(
    records: list[dict[str, Any]],
    column_map: dict[str, str],
    constants: dict[str, Any],
    *,
    stop_at_repeated_header: bool = True,
) -> list[dict[str, Any]]:
    """Apply *column_map*, attach *constants*, drop blank-PK rows.

    The first key in *column_map* is the primary-key field; rows with a blank PK
    are dropped.

    A row whose PK cell repeats the PK column's header string marks the start of a
    **different table** on the same sheet, not a cosmetic sub-header. Damodaran's
    ``ERPs by country`` sheet is the case that matters: below the Moody's-rated
    table it publishes a second table of unrated countries scored by PRS, whose
    columns do not line up (its column B is a risk score where the first table has
    the region, and its ERP sits where the first table has the default spread).
    Continuing past that boundary stored a wrong equity risk premium for 21
    countries. So parsing stops there, and how many rows were discarded is logged
    rather than passed over in silence.

    Set ``stop_at_repeated_header=False`` for a sheet where a repeated header is
    genuinely cosmetic and real data follows.
    """
    pk_field = next(iter(column_map))
    pk_xls_col = column_map[pk_field]  # e.g. "Country" or "Industry Name"
    out: list[dict[str, Any]] = []

    for index, record in enumerate(records):
        normalized: dict[str, Any] = dict(constants)
        for db_col, xls_col in column_map.items():
            if xls_col not in record:
                continue
            # Leave the PK column as-is (always a string identifier).
            if db_col == pk_field:
                raw = record[xls_col]
                normalized[db_col] = str(raw).strip() if raw is not None else None
            else:
                normalized[db_col] = _coerce_value(record[xls_col])

        pk_val = normalized.get(pk_field)
        if pk_val is None or (isinstance(pk_val, str) and pk_val == ""):
            continue
        # A repeated header cell (e.g. a row whose country cell literally reads
        # "Country") is the start of a second, differently-shaped table.
        if isinstance(pk_val, str) and pk_val == pk_xls_col:
            if stop_at_repeated_header:
                discarded = len(records) - index - 1
                log.info(
                    "damodaran.second_table.truncated",
                    pk_field=pk_field,
                    kept=len(out),
                    discarded=discarded,
                )
                break
            continue
        out.append(normalized)

    return out


# ---------------------------------------------------------------------------
# Derived columns and out-of-table scalars
# ---------------------------------------------------------------------------

#: Label of the pre-header cell of wacc.xls carrying the long-term government bond
#: rate. It is an input to Damodaran's sheet, not a column, so it needs its own
#: extraction path.
RISK_FREE_RATE_LABEL = "Long Term Treasury bond rate ="

#: Sheet of ctryprem.xls carrying per-country corporate tax rates. The main
#: "ERPs by country" sheet has no tax column.
COUNTRY_TAX_SHEET = "Country Tax Rates"


def derive_industry_columns(row: dict[str, Any]) -> dict[str, Any]:
    """Return ``row`` with the industry columns the published file only implies.

    ``wacc.xls`` publishes ``D/(D+E)``, not debt-to-equity, and no unlevered beta.
    Both are what the valuator actually needs: ``debt_to_equity`` is the only input
    from which ``_resolve_weights`` can build the equity/debt split, without which
    every assumption bundle fails to project onto the DCF.

    - ``debt_to_equity = w / (1 - w)`` where ``w = D/(D+E)``. ``w == 1`` (zero
      equity) leaves it ``None``: the ratio is undefined, and a non-finite double
      must never reach the DB.
    - ``beta_unlevered = beta_levered / (1 + (1 - tax_rate) * D/E)`` (Hamada).

    Pure: the input is not mutated. The ``debt_weight_raw`` parsing artefact is
    dropped from the result.
    """
    out = dict(row)
    weight = out.pop("debt_weight_raw", None)

    debt_to_equity: float | None = None
    if isinstance(weight, (int, float)) and 0.0 <= float(weight) < 1.0:
        w = float(weight)
        debt_to_equity = w / (1.0 - w)
    if weight is not None:
        # The source column was present: record the outcome explicitly, ``None``
        # included, so the column is written (as NULL) rather than silently absent.
        out["debt_to_equity"] = debt_to_equity

    beta_levered = out.get("beta_levered")
    if debt_to_equity is not None and isinstance(beta_levered, (int, float)):
        tax_rate = out.get("tax_rate")
        t = float(tax_rate) if isinstance(tax_rate, (int, float)) else 0.0
        out["beta_unlevered"] = float(beta_levered) / (1.0 + (1.0 - t) * debt_to_equity)

    return out


def parse_country_tax_rates(path: Path) -> dict[str, float]:
    """Country → corporate tax rate from the ``Country Tax Rates`` sheet.

    The sheet's header sits on row 0 with ``Country`` / ``Tax Rate`` in the first two
    columns (a second, unrelated country/year block sits further right and is
    ignored). Rows whose rate is not a fraction in ``[0, 1]`` are skipped rather
    than trusted.
    """
    rows, _sheet = _load_rows(path, COUNTRY_TAX_SHEET)
    out: dict[str, float] = {}
    for raw in rows[1:]:
        if len(raw) < 2:
            continue
        country = raw[0]
        rate = _coerce_value(raw[1])
        if not isinstance(country, str) or not country.strip():
            continue
        if not isinstance(rate, (int, float)):
            continue
        value = float(rate)
        if not 0.0 <= value <= 1.0:
            continue
        out[country.strip()] = value
    return out


def parse_preheader_scalar(path: Path, sheet_name: str, label: str) -> float | None:
    """Scalar input stored in a pre-header cell, found by its adjacent label.

    Damodaran's sheets carry inputs (the risk-free rate, the mature-market ERP)
    above the data header as ``label``/``value`` cell pairs rather than as columns.
    Scans for a cell whose text starts with ``label`` and returns the first numeric
    cell to its right on the same row. ``None`` when the label is absent.
    """
    rows, _sheet = _load_rows(path, sheet_name)
    needle = label.strip().lower()
    for raw in rows:
        for index, cell in enumerate(raw):
            if not isinstance(cell, str) or not cell.strip().lower().startswith(needle):
                continue
            for candidate in raw[index + 1 :]:
                value = _coerce_value(candidate)
                if isinstance(value, (int, float)):
                    return float(value)
    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_industry_xls(
    path: Path,
    *,
    region: str,
    year: int,
    column_map: dict[str, str],
    sheet_name: str | None = None,
    sheet_keywords: tuple[str, ...] | None = None,
) -> list[dict[str, Any]]:
    """Parse a Damodaran industry-level xls into normalized rows.

    Args:
        path: Path to the ``.xls`` / ``.xlsx`` file.
        region: Region tag injected into every row (e.g. ``"US"``).
        year: Data year injected into every row.
        column_map: Mapping ``db_column_name -> xls_header_string``.
        sheet_name: Explicit sheet to open; auto-detected when *None*.
        sheet_keywords: Keywords driving that auto-detection, most specific
            first; the module default is used when *None*.

    Returns:
        List of dicts ready for DB upsert.
    """
    records = _load_to_records(path, sheet_name, sheet_keywords)
    if not records:
        log.warning("damodaran.industry.empty_file", path=str(path))
        return []
    rows = _to_normalized_rows(records, column_map, {"region": region, "year": year})
    return [_null_non_numeric_industry_cells(row, column_map, path) for row in rows]


def parse_country_xls(
    path: Path,
    *,
    year: int,
    column_map: dict[str, str],
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Parse a Damodaran country-level xls into normalized rows.

    Args:
        path: Path to the ``.xls`` / ``.xlsx`` file.
        year: Data year injected into every row.
        column_map: Mapping ``db_column_name -> xls_header_string``.
        sheet_name: Explicit sheet to open; auto-detected when *None*.

    Returns:
        List of dicts ready for DB upsert.
    """
    records = _load_to_records(path, sheet_name)
    if not records:
        log.warning("damodaran.country.empty_file", path=str(path))
        return []
    return _to_normalized_rows(records, column_map, {"year": year})


# ---------- Downloader ----------


def download_dataset(url: str, dest: Path, timeout: float = 60.0) -> Path:
    """Download a Damodaran xls/xlsx file to dest. Overwrites if present."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        response = client.get(url)
        response.raise_for_status()
        dest.write_bytes(response.content)
    log.info("damodaran.download.ok", url=url, dest=str(dest), bytes=len(response.content))
    return dest


# ---------- Upserts ----------


def upsert_industry_rows(conn: duckdb.DuckDBPyConnection, rows: list[dict[str, Any]]) -> int:
    """Upsert into damodaran_industry for the (region, year) tuples present.

    Must be called inside an open transaction; does not manage its own BEGIN/COMMIT.
    """
    if not rows:
        return 0
    all_columns: set[str] = set()
    for r in rows:
        all_columns.update(r.keys())
    cols = sorted(all_columns)

    pairs = {(r["region"], r["year"]) for r in rows}
    for region, year in pairs:
        conn.execute(
            "DELETE FROM damodaran_industry WHERE region = ? AND year = ?",
            [region, year],
        )
    placeholders = ", ".join(["?"] * len(cols))
    col_list = ", ".join(cols)
    for r in rows:
        conn.execute(
            f"INSERT INTO damodaran_industry ({col_list}) VALUES ({placeholders})",
            [r.get(c) for c in cols],
        )
    return len(rows)


def upsert_country_rows(conn: duckdb.DuckDBPyConnection, rows: list[dict[str, Any]]) -> int:
    """Upsert into damodaran_country for the years present.

    Must be called inside an open transaction; does not manage its own BEGIN/COMMIT.
    """
    if not rows:
        return 0
    all_columns: set[str] = set()
    for r in rows:
        all_columns.update(r.keys())
    cols = sorted(all_columns)

    years = {r["year"] for r in rows}
    for year in years:
        conn.execute("DELETE FROM damodaran_country WHERE year = ?", [year])
    placeholders = ", ".join(["?"] * len(cols))
    col_list = ", ".join(cols)
    for r in rows:
        conn.execute(
            f"INSERT INTO damodaran_country ({col_list}) VALUES ({placeholders})",
            [r.get(c) for c in cols],
        )
    return len(rows)


# ---------- High-level importers ----------


def _import_files_into_run(
    conn: duckdb.DuckDBPyConnection,
    run: RefreshOutcome,
    *,
    industry_path: Path,
    country_path: Path,
    region: str,
    year: int,
    extra_industry_paths: dict[str, Path] | None = None,
    unavailable_datasets: Sequence[str] = (),
) -> None:
    """Parse + upsert both datasets, recording counts/status on ``run``.

    Shared body for :func:`import_damodaran_from_files` and
    :func:`import_damodaran` so a single :func:`refresh_run` envelope wraps the
    whole operation — exactly one refresh_log row per public call.

    ``extra_industry_paths`` maps :data:`INDUSTRY_DATASETS` keys to already-fetched
    files carrying the industry columns ``wacc.xls`` does not publish (margins,
    sales-to-capital, the multiples). ``unavailable_datasets`` names datasets the
    caller could not fetch at all; both a missing download and a dataset that
    fails to parse degrade the run to ``partial`` instead of aborting it (§13.2).
    """
    industry_parts: list[list[dict[str, Any]]] = [
        parse_industry_xls(
            industry_path,
            region=region,
            year=year,
            column_map=_WACC_DATASET.column_map,
            sheet_keywords=_WACC_DATASET.sheet_keywords,
        )
    ]
    degraded: list[str] = list(unavailable_datasets)
    for dataset in _EXTRA_DATASETS:
        path = (extra_industry_paths or {}).get(dataset.key)
        if path is None:
            continue
        try:
            part = parse_industry_xls(
                path,
                region=region,
                year=year,
                column_map=dataset.column_map,
                sheet_keywords=dataset.sheet_keywords,
            )
        except Exception as exc:
            # One unreadable file must not cost us the other eight columns.
            log.warning(
                "damodaran.dataset.parse_failed",
                dataset=dataset.key,
                path=str(path),
                error=str(exc),
            )
            degraded.append(dataset.key)
            continue
        if not part:
            log.warning("damodaran.dataset.no_rows", dataset=dataset.key, path=str(path))
        industry_parts.append(part)

    # Outer join on the industry label; the cost-of-capital file comes first and
    # so wins every column it shares with a later dataset.
    industry_rows = merge_industry_datasets(industry_parts)

    country_rows = parse_country_xls(
        country_path,
        year=year,
        column_map=DEFAULT_COUNTRY_COLUMN_MAP,
    )

    # Columns the published files only imply (D/(D+E) -> D/E, Hamada beta).
    industry_rows = [derive_industry_columns(row) for row in industry_rows]

    # Two country columns live outside the main sheet: the tax rate on its own
    # sheet, the risk-free rate in a pre-header input cell. Both degrade to a
    # warning if a future release moves or drops them (spec §13.2).
    try:
        tax_rates = parse_country_tax_rates(country_path)
    except (KeyError, ValueError) as exc:
        log.warning("damodaran.country_tax_sheet.unavailable", error=str(exc))
        tax_rates = {}
    try:
        risk_free_rate = parse_preheader_scalar(
            industry_path, "Industry Averages", RISK_FREE_RATE_LABEL
        )
    except (KeyError, ValueError) as exc:
        log.warning("damodaran.risk_free_rate.unavailable", error=str(exc))
        risk_free_rate = None
    for row in country_rows:
        country = row.get("country")
        if isinstance(country, str) and country in tax_rates:
            row["tax_rate"] = tax_rates[country]
        if risk_free_rate is not None:
            row["risk_free_rate"] = risk_free_rate

    # Detect empty-row situation before touching the DB.
    empty_sides: list[str] = []
    if not industry_rows:
        empty_sides.append("industry")
    if not country_rows:
        empty_sides.append("country")
    if empty_sides:
        log.warning(
            "damodaran.import.empty_rows",
            industry=len(industry_rows),
            country=len(country_rows),
            region=region,
            year=year,
        )

    with transaction(conn):
        ind = upsert_industry_rows(conn, industry_rows)
        ctry = upsert_country_rows(conn, country_rows)

    run.rows_affected = ind + ctry
    run.details = {
        "industry_rows": len(industry_rows),
        "country_rows": len(country_rows),
        "region": region,
        "year": year,
    }
    if degraded:
        run.details["unavailable_datasets"] = sorted(degraded)
    reasons: list[str] = []
    if empty_sides:
        reasons.append(f"empty rows for: {', '.join(empty_sides)}")
    if degraded:
        reasons.append(f"datasets unavailable: {', '.join(sorted(degraded))}")
    if reasons:
        run.status_override = "partial"
        run.error_message_override = "; ".join(reasons)


def import_damodaran_from_files(
    conn: duckdb.DuckDBPyConnection,
    *,
    industry_path: Path,
    country_path: Path,
    region: str,
    year: int,
    extra_industry_paths: dict[str, Path] | None = None,
) -> IngestResult:
    """Import already-downloaded Damodaran files into the DB.

    Args:
        extra_industry_paths: Optional :data:`INDUSTRY_DATASETS` key → file for the
            additional industry datasets (margins, sales-to-capital, multiples).
            Omit it to import the cost-of-capital and country files alone.
    """
    with refresh_run(
        conn,
        source="damodaran",
        log=log,
        error_event="damodaran.import.failed",
        log_fail_event="damodaran.refresh_log_insert_failed",
    ) as run:
        _import_files_into_run(
            conn,
            run,
            industry_path=industry_path,
            country_path=country_path,
            region=region,
            year=year,
            extra_industry_paths=extra_industry_paths,
        )
    assert run.result is not None  # refresh_run always sets it on exit
    return run.result


def import_damodaran(
    conn: duckdb.DuckDBPyConnection,
    *,
    download_dir: Path,
    region: str = "US",
    year: int | None = None,
    industry_url: str = INDUSTRY_WACC_URL,
    country_url: str = COUNTRY_RISK_URL,
) -> IngestResult:
    """Download and import current-year Damodaran datasets.

    Every dataset in :data:`INDUSTRY_DATASETS` is fetched, not just ``wacc.xls``:
    the margin / capex / multiple files carry the industry columns the valuator
    and the screener need. Download and import share a single :func:`refresh_run`
    envelope, so the run is recorded in refresh_log exactly once. Failing to fetch
    the cost-of-capital or country file surfaces as an error run carrying the
    legacy ``download failed: ...`` message; failing to fetch one of the extra
    industry files only degrades the run to ``partial`` (§13.2).
    """
    year = year or datetime.now().year
    with refresh_run(
        conn,
        source="damodaran",
        log=log,
        error_event="damodaran.refresh.failed",
        log_fail_event="damodaran.refresh_log_insert_failed",
    ) as run:
        industry_path: Path | None = None
        country_path: Path | None = None
        extra_paths: dict[str, Path] = {}
        unavailable: list[str] = []
        try:
            industry_path = download_dataset(industry_url, download_dir / "wacc.xls")
            country_path = download_dataset(country_url, download_dir / "ctryprem.xls")
            for dataset in _EXTRA_DATASETS:
                try:
                    extra_paths[dataset.key] = download_dataset(
                        dataset.url, download_dir / f"{dataset.key}.xls"
                    )
                except Exception as exc:
                    # A single missing file costs only its own columns.
                    log.warning(
                        "damodaran.dataset.download_failed",
                        dataset=dataset.key,
                        url=dataset.url,
                        error=str(exc),
                    )
                    unavailable.append(dataset.key)
        except Exception as e:
            # Record the run as a download error with the legacy message; the
            # status_override keeps it a single refresh_log row and lets the
            # specific download event be logged here (not the generic envelope).
            run.status_override = "error"
            run.error_message_override = f"download failed: {e}"
            log.exception("damodaran.download.failed", error=str(e))
        if industry_path is not None and country_path is not None:
            try:
                _import_files_into_run(
                    conn,
                    run,
                    industry_path=industry_path,
                    country_path=country_path,
                    region=region,
                    year=year,
                    extra_industry_paths=extra_paths,
                    unavailable_datasets=unavailable,
                )
            except Exception as e:
                # A parse/upsert failure is an import error, not a download one;
                # log it under the import event the from-files path also uses.
                run.status_override = "error"
                run.error_message_override = str(e)
                log.exception("damodaran.import.failed", error=str(e))
    assert run.result is not None  # refresh_run always sets it on exit
    return run.result
